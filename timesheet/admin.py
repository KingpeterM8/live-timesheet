import calendar
import datetime
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import timedelta, date
from collections import defaultdict
from django.utils.html import format_html
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.admin import GroupAdmin as BaseGroupAdmin
from django.contrib.auth.models import User, Group
from django.urls import reverse
from django.shortcuts import redirect

# Unfold imports
from unfold.admin import ModelAdmin
from unfold.decorators import action

from .models import TimesheetEntry, TradeAllocation

admin.site.unregister(User)
admin.site.unregister(Group)


@admin.register(User)
class UserAdmin(BaseUserAdmin, ModelAdmin):
    # Adds the button to the 'Actions' menu on the user's edit page
    actions_detail = ["change_password_custom"]

    @action(description="Change User Password", url_path="change-password")
    def change_password_custom(self, request, *args, **kwargs):
        # 1. Grab the ID from the URL arguments (passed as 'object_id')
        user_id = kwargs.get("object_id")

        # 2. Redirect to the built-in password change form for that ID
        return redirect(reverse("admin:auth_user_password_change", args=[user_id]))

@admin.register(Group)
class GroupAdmin(BaseGroupAdmin, ModelAdmin):
    pass

@admin.register(TimesheetEntry)
class TimesheetEntryAdmin(ModelAdmin):
    # Use full_name instead of user in the list display
    list_display = (
        "display_full_name",
        "date_worked",
        "start_time",
        "end_time",
        "hours_worked",
        "status_tag"
    )

    def display_full_name(self, obj):
        first = obj.user.first_name
        last = obj.user.last_name
        full_name = f"{first} {last}".strip() or obj.user.username

        return format_html(
            '<span class="font-semibold text-gray-900 dark:text-gray-100">{}</span>',
            full_name
        )

    display_full_name.short_description = "Employee"
    # CRITICAL: This allows the 'ordering' attribute above to work for this column
    display_full_name.admin_order_field = "user__first_name"

    # --- Status Badge (same as before) ---
    def status_tag(self, obj):
        colors = {
            "APPROVED": "bg-green-100 text-green-700 border-green-200",
            "PENDING": "bg-orange-100 text-orange-700 border-orange-200",
            "REJECTED": "bg-red-100 text-red-700 border-red-200",
        }
        color_class = colors.get(obj.status, "bg-gray-100 text-gray-700 border-gray-200")
        return format_html(
            '<span class="px-2 py-1 text-xs font-semibold border rounded-full {}">{}</span>',
            color_class,
            obj.status
        )

    status_tag.short_description = "Status"

    # 1. Bulk Approval Action
    @action(description="Approve Selected Timesheets", icon="check_circle")
    def approve_timesheets(self, request, queryset):
        queryset.update(status='APPROVED')
        self.message_user(request, "Selected timesheets have been approved.")

    # 2. Excel Export Action
    @action(description="Export Bi-Weekly Payroll", icon="description")
    def export_to_excel(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, "No records selected.", level='warning')
            return

        # Period Detection
        latest_entry_date = queryset.order_by('-date_worked').first().date_worked
        year, month = latest_entry_date.year, latest_entry_date.month
        if latest_entry_date.day <= 15:
            start_date, end_date = date(year, month, 1), date(year, month, 15)
        else:
            start_date = date(year, month, 16)
            end_date = date(year, month, calendar.monthrange(year, month)[1])

        wb = Workbook()
        wb.remove(wb.active)

        # Styles
        bold_font = Font(bold=True)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        data_by_user = defaultdict(list)
        period_queryset = queryset.filter(date_worked__range=(start_date, end_date)).order_by('date_worked')
        for entry in period_queryset:
            data_by_user[entry.user].append(entry)

        for user, entries in data_by_user.items():
            ws = wb.create_sheet(title=str(user.username)[:31])

            # Row 1 height and narrow buffer columns
            ws.row_dimensions[1].height = 15
            ws.column_dimensions['A'].width = 3.00
            ws.column_dimensions['J'].width = 3.00
            ws.column_dimensions['K'].width = 3.00

            # --- Header Section ---
            ws.merge_cells('B2:H2')
            ws['B2'] = "Empyreal Painting & Construction"
            ws['B2'].font = Font(size=14, bold=True)
            ws.merge_cells('B3:H3')
            ws['B3'] = '"Built on trust, focused on quality."'
            ws['B3'].font = Font(size=11, italic=True)

            ws['B5'], ws['C5'] = "Employee Name:", f"{user.first_name} {user.last_name}"
            ws['B6'] = "Pay Period:"
            ws.merge_cells('C6:D6')
            ws['C6'] = f"{start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"
            ws['B5'].font = ws['B6'].font = bold_font
            ws['C6'].alignment = left_aligned

            # --- Table Headers ---
            headers = ['Day of Week', 'Date', 'Job', 'In', 'Lunch Start', 'Lunch End', 'Out', 'Hours']
            for col_num, text in enumerate(headers, 2):
                cell = ws.cell(row=8, column=col_num, value=text)
                cell.font, cell.fill, cell.border, cell.alignment = header_font, header_fill, thin_border, center_aligned

            # Week Cutoff Row
            ws.merge_cells('B16:I16')
            ws['B16'] = "---Week Cutoff---"
            ws['B16'].font, ws['B16'].alignment = bold_font, center_aligned

            entries_dict = {e.date_worked: e for e in entries}
            current_date = start_date
            row_num = 9
            max_job_length = 0

            while current_date <= end_date:
                if row_num == 16:
                    row_num += 1

                entry = entries_dict.get(current_date)
                job_address = entry.task_description if entry else ""
                if job_address:
                    max_job_length = max(max_job_length, len(str(job_address)))

                # Lunch End calculation
                lunch_end_str = ""
                if entry and entry.lunch_start_time and hasattr(entry, 'lunch_duration'):
                    dummy = datetime.date.today()
                    start_dt = datetime.datetime.combine(dummy, entry.lunch_start_time)
                    end_dt = start_dt + datetime.timedelta(minutes=entry.lunch_duration)
                    lunch_end_str = end_dt.strftime('%I:%M %p')

                row_data = [
                    current_date.strftime('%A'),
                    current_date.strftime('%m/%d/%Y'),
                    job_address,
                    entry.start_time.strftime('%I:%M %p') if entry and entry.start_time else "",
                    entry.lunch_start_time.strftime('%I:%M %p') if entry and entry.lunch_start_time else "",
                    lunch_end_str,
                    entry.end_time.strftime('%I:%M %p') if entry and entry.end_time else "",
                    float(entry.hours_worked) if entry and entry.hours_worked else None
                ]

                is_stripe = (row_num % 2 == 0)
                for col_offset, value in enumerate(row_data, 2):
                    cell = ws.cell(row=row_num, column=col_offset, value=value)
                    cell.border, cell.alignment = thin_border, center_aligned
                    if is_stripe: cell.fill = stripe_fill
                    if col_offset == 9 and value is not None: cell.number_format = '0.00'

                current_date += timedelta(days=1)
                row_num += 1

            # Main Total
            ws.cell(row=row_num, column=8, value="TOTAL:").font = bold_font
            sum_cell = ws.cell(row=row_num, column=9, value=f"=SUM(I9:I{row_num - 1})")
            sum_cell.font, sum_cell.border, sum_cell.alignment, sum_cell.number_format = bold_font, thin_border, center_aligned, '0.00'

            # --- Side-Panel L&I Trade Matrix ---
            ws.merge_cells('L6:M6')
            ws['L6'] = "L&I Trade Matrix"
            ws['L6'].font = bold_font
            for col, val in [(12, "Trade Type"), (13, "Total Hours")]:
                cell = ws.cell(row=8, column=col, value=val)
                cell.font, cell.fill, cell.border, cell.alignment = header_font, header_fill, thin_border, center_aligned

            trades = ['Painting - Interior', 'Painting - Exterior', 'Roofing', 'Framing', 'Finish Carpentry',
                      'Flooring', 'Fencing', 'Remodel / Repair', 'Estimator']
            trade_totals = defaultdict(float)
            for entry in entries:
                for alloc in TradeAllocation.objects.filter(timesheet=entry):
                    trade_totals[alloc.trade_type] += float(alloc.hours_allocated)

            matrix_row = 9
            for trade in trades:
                is_stripe = (matrix_row % 2 == 0)
                cl = ws.cell(row=matrix_row, column=12, value=trade)
                cm = ws.cell(row=matrix_row, column=13, value=trade_totals.get(trade, 0) or "")
                for c in [cl, cm]:
                    c.border = thin_border
                    if is_stripe: c.fill = stripe_fill
                    if c.column == 13: c.alignment, c.number_format = center_aligned, '0.00'
                matrix_row += 1

            ws.cell(row=18, column=12, value="TOTAL:").font = bold_font
            msum = ws.cell(row=18, column=13, value="=SUM(M9:M17)")
            msum.font, msum.border, msum.alignment, msum.number_format = bold_font, thin_border, center_aligned, '0.00'

            # Sizing
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = max_job_length + 4 if max_job_length > 12 else 12
            for col in ['E', 'F', 'G', 'H', 'I']: ws.column_dimensions[col].width = 12
            ws.column_dimensions['L'].width = 20
            ws.column_dimensions['M'].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Empyreal_Payroll_{end_date}.xlsx"'
        wb.save(response)
        return response

    def get_actions(self, request):
        """Remove the default 'delete_selected' action for safety."""
        actions = super().get_actions(request)
        if 'delete_selected' in actions:
            del actions['delete_selected']
        return actions

    def has_delete_permission(self, request, obj=None):
        """Disable deletion entirely for this model in the admin."""
        return False

    # Assign actions to the class
    actions = [export_to_excel, approve_timesheets]


# Keeping your Inline as is
class TradeAllocationInline(admin.TabularInline):
    model = TradeAllocation
    extra = 0

